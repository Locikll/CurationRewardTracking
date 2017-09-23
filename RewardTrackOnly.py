"""
This bot tracks the votes of Curie's "Vote Followers", and logs them after payout with the Curation rewards.

Written by @Locikll

Requires modules:
piston-lib,steem (pip install steem), openpyxl (pip install openpyxl)

"""
import sys
import datetime
import os
import subprocess
import math
import re
import csv
import time

from openpyxl import *
import pickle

from time import gmtime, strftime
import os.path
from pathlib import Path
import timeit

import multiprocessing.dummy as mp 
from multiprocessing import process
from collections import OrderedDict

from datetime import datetime, timedelta

import piston
import steem as pysteem
from piston.steem import Steem
from random import randint

#EDITABLE VARIABLES {{

#Steemit Node & Account stuff
Node = "wss://this.piston.rocks"

steemPostingKey = ''  #Private posting key
steemAccountName = 'curie' 

#List of Followed Curators
followedcurators = ['frontpage','blitz','eureka','homesteadbuilder','poeticammo','dna-replication','archdruid','skadi','zeks','camelot','thunderbird','teacherspet']


#Recent Activity limit (Increase this if users have large number of activities / second), 5 should be a large enough size
ACTLIMIT = 5


# Number of Seconds of following votes between checking the votes to see if any correspond to a curation reward (ie 1800 means 30 minutes of following votes and 1 iteration of checking the votes)
seccheck = 1800
# Curation reward history limit (at the current state of 1500 this goes back 3 days of Curation-rewards)
REWHLIM = 1500
#VOTER and Reward History account
#VOTER = 'locikll'
RewHistACC = 'curie'

#Number of processing threads to allocate for multithreading purposes
PROCESSINGTHREADS = 4


#EDITABLE VARIABLES STOP/END  }}


steem = Steem(wif=steemPostingKey,node=Node)

if Path('curatordict.pickle').is_file():
    curatordict = pickle.load(open('curatordict.pickle','rb'))
else:
    curatordict = OrderedDict((el,[]) for el in followedcurators)

#print(curatordict)

#Initial File Manipulation and storage
wb = Workbook()
ws = wb.active
filepath = 'reports'

def setupfiledir():
 
    if not os.path.exists(filepath):
        os.makedirs(filepath)
    
    for Files in range(0,len(followedcurators)):
        
        filenamepath = filepath+'/'+followedcurators[Files]+'.xlsx'
        curfiles = Path(filenamepath)
        
        if not curfiles.is_file():
            
            ws['A1'] = 'Post ID'
            ws['B1'] = 'Date'
            ws['C1'] = 'Curation Reward (STEEM)'
            ws['D1'] = 'Running Curation Reward Total (STEEM)'
            ws['E1'] = 'Total Reward (STEEM)'
            
            wb.save(filename=filenamepath)

#Run initial function for setting up directories/checking files            
setupfiledir()            
    
#Check users history for votes for following and tracking votes               
def votefeed(usrn):
    
    curatoraccs = list(curatordict.keys())[usrn]
    
    try:
        Curatorvotes = list(steem.get_account_history(curatoraccs, limit=ACTLIMIT))
    except:
        Curatorvotes = []
        print('Account history Exception, trying again')
        pass
    
    #print(Curatorvotes)
    for checkpost in range(0,len(Curatorvotes)):
        
        
        try:
            Isvote = Curatorvotes[checkpost][1]['op'][0]
        except Exception:
            print(Exception)
            print('Isvote Error')
            Isvote = ''
            pass        
         
        if Isvote=='vote':
            
            votername = Curatorvotes[checkpost][1]['op'][1]['voter']
            
            userweight = Curatorvotes[checkpost][1]['op'][1]['weight'] / 100
            
            permlink = Curatorvotes[checkpost][1]['op'][1]['permlink']
            author = Curatorvotes[checkpost][1]['op'][1]['author']
            identifier = '@'+author+'/'+permlink

            postdata = get_post(identifier)
            
            postid = postdata[0]
            ptitle = postdata[1]
            activevotes = postdata[2]
            pauthor = postdata[3]
            
            IDIN = []
            
            for chkID in range(0,len(followedcurators)):    
                
                IDIN.append( any(e[0] == identifier for e in curatordict[followedcurators[chkID]]) )
                
            #Make sure the votername is the same as the curator
            if True not in IDIN and votername == curatoraccs:

                votetime = (list(filter(lambda voter: voter['voter']==curatoraccs,steem.get_post(identifier).active_votes))[0]['time']).replace('T',' ')
            
                curatordict[curatoraccs].append([identifier,votetime])

                print('Post has been voted on: '+str(identifier))

                              
        else:
            continue
         
    #Save Data at Each point so that if program shuts down, it can be loaded again with the previous state
    pickle_out = open('curatordict.pickle','wb')
    pickle.dump(curatordict,pickle_out)
    pickle_out.close()
    

    
#Check the post rewards to see whether the votes are within the past 2-3 days of voting rewards.    
def checkrewards(curname):
    
    if len(curatordict[followedcurators[curname]]) > 0:
        
        Rewardhistory = list(pysteem.account.Account(RewHistACC).get_account_history(filter_by='curation_reward',limit=REWHLIM,index=-1,order=1))
    
        Curatorsvotedposts = curatordict[followedcurators[curname]]
        
        indel = []
        
        for Post in range(0,len(Curatorsvotedposts)):
            
            POSTID = Curatorsvotedposts[Post][0]
            VOTETIME = Curatorsvotedposts[Post][1]
            
            
            votetimedatetime = datetime.strptime(VOTETIME,'%Y-%m-%d %H:%M:%S')
            dayssincevote = (datetime.utcnow() - votetimedatetime).seconds / (3600*24)
       

            for rewardpost in range(0,len(Rewardhistory)):
                reperm = Rewardhistory[rewardpost]['comment_permlink']
                reauth = Rewardhistory[rewardpost]['comment_author']
                reiden = '@'+reauth+'/'+reperm
                
                
                if reiden == POSTID:
                    rewardMvest = float((Rewardhistory[rewardpost]['reward'].split()[0])) / 1e6
                    
                    #Calculate Steem per mvests
                    steemperMvest = pysteem.account.Account('curie').converter.steem_per_mvests()
                    
                    steemreward = rewardMvest * steemperMvest
                    
                    curatordict[followedcurators[curname]][Post].append(steemreward)
                    
                    wb = load_workbook(filepath+'/'+followedcurators[curname]+'.xlsx')
                    ws = wb.active

                    ws.append(curatordict[followedcurators[curname]][Post])
                    
                    lastrow = ws.max_row
                    
                    ws[("D"+str(lastrow))] = "=SUM(C1:"+"C"+str(lastrow)+")"           
                    ws["E2"] = "=SUM(C:C)"
                    
                    wb.save(filepath+'/'+followedcurators[curname]+'.xlsx')

                    print(POSTID)
                    
                    #Indexes of post to delete from Curatordict to clear up memory as it has already been logged
                    indel.append(Post)
                            
        #DELETE from Curatordict to clear memory
        for dkey in range(0,len(indel)):
            # -dkey so that when an item is removed, it keeps the index in range
            IDXDELETE = indel[dkey] - dkey
            del(curatordict[followedcurators[curname]][IDXDELETE])
    
            #print(dayssincevote)

def get_post(identifier):
    
    try:
        postid = steem.get_post(identifier)
        posttitle = postid.title        
        postvotes = postid.active_votes
        postauthor = postid.author
        
    except:
        print('Exception occured with Identifier: '+identifier)
        postid = ''
        posttitle = ''
        postvotes = ''
        postauthor = ''
        pass
        
    return [postid,posttitle,postvotes,postauthor]        



if __name__ == "__main__":
    n = 0
    while True:
        try:            
            p=mp.Pool(PROCESSINGTHREADS)
            p.map(votefeed,range(0,len(followedcurators)))
            p.close()
            p.join()
            clk = time.clock() 
            
            if (clk-n*seccheck)>seccheck:
                print(clk-n*seccheck)
                q=mp.Pool(PROCESSINGTHREADS)
                q.map(checkrewards,range(0,len(followedcurators)))    
                q.close()
                q.join()
                n=n+1
            
        except (KeyboardInterrupt):
            print("Quitting...")
            break
        except Exception as e:
            print("### Exception Occurred: Restarting...")



            
